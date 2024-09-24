# openpyxl - criando uma planilha do Excel (Workbook e Worksheet)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
worksheet: Worksheet = workbook.active

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')
worksheet.cell(1, 4, 'Turma')
students = [
    # nome      idade nota turma
    ['João',    14,   5.5, 'A'],
    ['Maria',   13,   9.7, 'B'],
    ['Luiz',    15,   8.8, 'C'],
    ['Alberto', 16,   10,  'D'],
    ['Davi',    14,   3.2, 'A'],
    ['José',    13,   5.7, 'B'],
    ['Carlos',  13,   1.8, 'C'],
    ['Luiza',   12,   4.7, 'C'],
    ['Rose',    17,   7.5, 'A'],
    ['Caio',    15,   9.0, 'D'],
    ['Cléo',    12,   8.3, 'B'],
    ['Alex',    16,   6.7, 'D'],
]
# Lógica importantissima de se entender:

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)
workbook.save(WORKBOOK_PATH)